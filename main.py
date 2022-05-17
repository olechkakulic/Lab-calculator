import pandas as pd
import time
import streamlit as st
import plotly.express as px
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from urllib.request import urlopen
import json
import os

from openpyxl import load_workbook
@st.cache
def load_dataset(data_link):
    dataset = pd.read_csv(data_link)
    return dataset


titanic_link = 'https://raw.githubusercontent.com/mwaskom/seaborn-data/master/titanic.csv'
titanic_data = load_dataset(titanic_link)
#ЗАГОЛОВКИ #
st.title("Калькулятор Брэгговского зеркала")
st.markdown("Люблю Дашеньку Салтыкову")
#код для вводимых данных
st.sidebar.number_input('Длина волны', min_value = 0, max_value = 1000, step = 1)
st.sidebar.number_input('Значение угла', min_value = 0, max_value = 180, step = 1)
st.sidebar.selectbox(
     'Тип поляризации',
     ('TE', 'TM',))
material = ['Ag',"Au"]
st.sidebar.selectbox(
     'Материал',
     (material))


# прога ептить пробная попытка запустить файл нахуй.
wb = load_workbook(r"C:\Users\4814285\lab culc\Johnson.xlsx")
sheet = wb.get_sheet_by_name('Worksheet')
# норм попытка для Ag
wb = load_workbook("D:\Ag.xlsx")
sheet = wb.get_sheet_by_name('Worksheet')
#b - Длина волны из столбцов
#def blocks(material,b)
    #if b != sheet.cell(column=1).value:
#return **
# работа с папкой, в которой лежат файлы
directory = 'D:\материалы'
files = os.listdir(directory)
for i in range (len(files)):
    input(files[i])
    os.path.splitext(D:\материалы)[0]



